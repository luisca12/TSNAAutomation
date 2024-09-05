import openpyxl
import traceback

from log import authLog

def openNDLM():

    dataList = []
    headerRow = None
    columnIndices = {}
    headers = ['Action', 'Hostname', 'SERIAL NUMBER']

    while True:
        try:
            file = input(f"Please enter the NDLM file name:")
            ndlmFile = openpyxl.load_workbook(file)
            ndlmFileSheet = ndlmFile.active
            break
        
        except FileNotFoundError:
            print(f"File: {file} not found. Please check the file name and try again, remember to include the .xlsx")
            authLog.error(f"File not found in path {file}")
            authLog.error(traceback.format_exc())

        except Exception as error:
            print(f"File: {file} not found. Please check the file name and try again, remember to include the .xlsx")
            authLog.error(f"Wasn't possible to choose the XLSX file: {file}, error message: {error}")
            authLog.error(traceback.format_exc())

    try:
        for row in ndlmFileSheet.iter_rows():
            if headerRow is None:
                for cell in row:
                    if cell.value in headers:
                        columnIndices[cell.value] = cell.column
                if len(columnIndices) == len(headers):
                    headerRow = row[0].row
                    break
        
        if columnIndices:
            for n, row in enumerate(ndlmFileSheet.iter_rows(min_row=headerRow + 1, values_only=True), start=1):
                actionValue = row[columnIndices['Action'] - 1]
                hostnameValue = row[columnIndices['Hostname']  - 1]
                serialValue = row[columnIndices['SERIAL NUMBER'] - 1]
            
                if actionValue and hostnameValue and serialValue:
                    action = f'Action{n}'
                    dataEntry = {
                        f'{action}' : actionValue,
                        'Details': {
                            'Hostname' : hostnameValue,
                            'Serial Number': serialValue
                        }
                    }
                    dataList.append(dataEntry)
        return dataList

    except Exception as error:
        print(f"ERROR: {error}\n", traceback.format_exc())
        authLog.error(f"Error while reading the xlsx file, error message: {error}\n", traceback.format_exc())

ndlmData = openNDLM()
for item in ndlmData:
    print(item)

def handleNDLM(actionList, hostnameList, serialList):
    actionList[0] # API calls
    hostnameList[0] # Replaces the hostname in the API payload
    serialList[0] # Replaces the serial number in the API payload

    pass